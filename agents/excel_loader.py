"""excel_loader.py — Parse FSLDM mapping spec Excel into MappingSpec + schemas.

Expected sheet layout (FSLDM_Deposit_Mapping_Spec_COMPLETE.xlsx style):
  COVER              meta (Mapping ID, Source/Target System)
  SOURCE_SCHEMA      Table | Column | Data Type | Nullable | PK | FK/Notes
  FCT_*              row 1 = title, row 2 = header, rows 3+ = mappings
                     Headers: Target Column | Type | Null | Transform |
                              Source Table(s) | Source Expression | Products | Notes
"""
from __future__ import annotations
from io import BytesIO
from typing import Any

from agents.state import FieldMapping, MappingSpec, TargetTable

CONFIDENCE_BY_TRANSFORM = {
    "direct": 0.95,
    "literal": 1.0,
    "lookup": 0.85,
    "derived": 0.7,
    "computed": 0.7,
    "join": 0.85,
    "case": 0.6,
    "etl": 1.0,
}


def _conf_for(transform: str | None, source_expr: str | None) -> float:
    t = (transform or "").lower().strip()
    if not (source_expr or "").strip():
        return 0.3
    for key, v in CONFIDENCE_BY_TRANSFORM.items():
        if key in t:
            return v
    return 0.7


def parse_excel(content: bytes, dialect: str = "teradata") -> tuple[dict, dict, MappingSpec]:
    """Returns (source_schema, target_schema, mapping_spec) from a mapping-spec workbook."""
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)

    # ── COVER ────────────────────────────────────────────────────────────
    mapping_id = "FSLDM-EXCEL-001"
    source_system = "CORE_BANKING_ODS"
    target_system = "EDW_FSLDM"
    if "COVER" in wb.sheetnames:
        for row in wb["COVER"].iter_rows(values_only=True):
            for i, cell in enumerate(row):
                if cell and isinstance(cell, str):
                    label = cell.strip().lower()
                    val = row[i + 1] if i + 1 < len(row) else None
                    if label == "mapping id" and val:
                        mapping_id = str(val).strip()
                    elif label == "source system" and val:
                        source_system = str(val).strip()
                    elif label == "target system" and val:
                        target_system = str(val).strip()

    # ── SOURCE_SCHEMA ────────────────────────────────────────────────────
    source_tables: dict[str, dict] = {}
    if "SOURCE_SCHEMA" in wb.sheetnames:
        ws = wb["SOURCE_SCHEMA"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            header = [str(c or "").strip().lower() for c in rows[0]]

            def col(name: str) -> int:
                for i, h in enumerate(header):
                    if name in h:
                        return i
                return -1

            ti, ci, dti, ni, pi, fi = col("table"), col("column"), col("type"), col("null"), col("pk"), col("fk")
            for r in rows[1:]:
                if not r or not r[ti] or not r[ci]:
                    continue
                tbl = str(r[ti]).strip()
                source_tables.setdefault(tbl, {"name": tbl, "columns": []})
                source_tables[tbl]["columns"].append({
                    "name": str(r[ci]).strip(),
                    "type": str(r[dti] or "").strip() if dti >= 0 else "",
                    "nullable": (str(r[ni] or "").strip().upper() != "N") if ni >= 0 else True,
                    "pk": (str(r[pi] or "").strip().upper() == "Y") if pi >= 0 else False,
                    "fk": str(r[fi] or "").strip() if fi >= 0 else "",
                })

    source_schema = {
        "system": source_system,
        "dialect": dialect,
        "tables": list(source_tables.values()),
    }

    # ── Per target FCT_* sheet ───────────────────────────────────────────
    target_tables: list[TargetTable] = []
    target_schema_tables: list[dict] = []
    fact_sheets = [s for s in wb.sheetnames if s.upper().startswith(("FCT_", "DIM_"))]

    for sheet_name in fact_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue

        # Find the header row (row containing "Target Column")
        header_row_idx = None
        for i, r in enumerate(rows[:5]):
            joined = " ".join(str(c or "") for c in r).lower()
            if "target column" in joined:
                header_row_idx = i
                break
        if header_row_idx is None:
            continue

        header = [str(c or "").strip().lower() for c in rows[header_row_idx]]

        def col(*names: str) -> int:
            for i, h in enumerate(header):
                if any(n in h for n in names):
                    return i
            return -1

        c_target = col("target column")
        c_type = col("type", "data type")
        c_null = col("null")
        c_transform = col("transform")
        c_src_tbl = col("source table")
        c_src_expr = col("source expression", "expression")
        c_notes = col("business rule", "note")

        fms: list[FieldMapping] = []
        opens: list[str] = []
        cols_meta: list[dict] = []

        for r in rows[header_row_idx + 1:]:
            if not r or not r[c_target]:
                continue
            tgt = str(r[c_target]).strip()
            if not tgt or tgt.lower().startswith("--"):
                continue

            transform = str(r[c_transform] or "").strip() if c_transform >= 0 else ""
            src_expr = str(r[c_src_expr] or "").strip() if c_src_expr >= 0 else ""
            src_tbl_str = str(r[c_src_tbl] or "").strip() if c_src_tbl >= 0 else ""
            notes = str(r[c_notes] or "").strip() if c_notes >= 0 else ""

            src_tables_list = [t.strip() for t in src_tbl_str.replace(",", "/").split("/") if t.strip()]
            conf = _conf_for(transform, src_expr)
            open_q = None
            if conf < 0.5 or not src_expr:
                open_q = f"Confirm lineage for {tgt}"
                opens.append(open_q)

            fms.append(FieldMapping(
                target_column=tgt,
                source_expr=src_expr or f"/* TODO: source for {tgt} */ NULL",
                source_tables=src_tables_list,
                transform_note=transform or notes or "",
                confidence=conf,
                open_question=open_q,
            ))

            cols_meta.append({
                "name": tgt,
                "type": str(r[c_type] or "").strip() if c_type >= 0 else "",
                "nullable": (str(r[c_null] or "").strip().upper() != "N") if c_null >= 0 else True,
            })

        # Title-line grain hint
        title = " ".join(str(c or "") for c in rows[0])
        grain = ""
        if "grain" in title.lower():
            grain = title.split("grain", 1)[-1].lstrip(": ").strip()
        if not grain:
            grain = sheet_name

        target_tables.append(TargetTable(
            target_table=sheet_name,
            grain_description=grain,
            field_mappings=fms,
            open_questions=opens[:6],
        ))
        target_schema_tables.append({
            "name": sheet_name,
            "grain": grain,
            "columns": cols_meta,
        })

    target_schema = {
        "system": target_system,
        "dialect": dialect,
        "tables": target_schema_tables,
    }

    spec = MappingSpec(
        mapping_id=mapping_id,
        dialect=dialect,
        target_tables=target_tables,
        notes=f"Parsed from Excel mapping spec. {len(fact_sheets)} target tables.",
    )

    return source_schema, target_schema, spec
